# -*- coding: utf-8 -*-
"""
working_output.xlsx yazıcısı.

İki sayfa:
  * 'Result'  -> her working satırı + işlem sonuç kolonları (✓/✗)
  * 'Ignored' -> IP/hostname uyuşmazlığı vb. nedenle atlanan satırlar + açıklama

✓ başarılı, ✗ sorunlu, - atlandı. ✓ yeşil, ✗ kırmızı; INFO'da ayrıntı.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common.logging_setup import get_logger
from migration.processor import FAIL, OK, RowResult

log = get_logger("output_writer")

RESULT_HEADERS = [
    "PamEnvanterSatır",
    "username",
    "ip address",
    "hostname",
    "OS",
    "safe name",
    "domain",
    "User Group",        # 8
    "User",              # 9
    "Member",            # 10
    "Managed System",    # 11
    "Managed Account",   # 12
    "Link",              # 13
    "Smart Rule",        # 14
    "AccessLevel",       # 15
    "Role",              # 16
    "Group ID",
    "User ID",
    "MS ID",
    "MA ID",
    "SR ID",
    "INFO",
]

IGNORED_HEADERS = [
    "PamEnvanterSatır",
    "username",
    "ip address",
    "hostname",
    "OS",
    "safe name",
    "domain",
    "INFO",
]

# Durum (✓/✗) kolonlarının indeksleri (1-based)
_MARK_COLS = (8, 9, 10, 11, 12, 13, 14, 15, 16)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_IGNORED_HEADER_FILL = PatternFill("solid", fgColor="C0504D")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_OK_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_header(ws, headers, fill) -> None:
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = _HEADER_ALIGN


def _autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 70)


def write_output(output_file: Path, results: List[RowResult],
                 sheet_name: str = "Result", ignored_sheet_name: str = "Ignored") -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Result sayfası ---
    ws = wb.active
    ws.title = sheet_name
    _write_header(ws, RESULT_HEADERS, _HEADER_FILL)
    for r in results:
        ws.append([
            r.pam_satir, r.username, r.ip_address, r.hostname, r.os, r.safe_name, r.domain,
            r.user_group_mark, r.user_mark, r.member_mark,
            r.managed_system_mark, r.managed_account_mark, r.link_mark,
            r.smart_rule_mark, r.access_level_mark, r.role_mark,
            r.user_group_id, r.user_id, r.managed_system_id, r.managed_account_id, r.smart_rule_id,
            r.info,
        ])
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in _MARK_COLS:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _CENTER
            if cell.value == OK:
                cell.fill = _OK_FILL
            elif cell.value == FAIL:
                cell.fill = _FAIL_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(RESULT_HEADERS))}{ws.max_row}"
    _autosize(ws)

    # --- Ignored sayfası ---
    ignored = [r for r in results if r.ignored]
    ws_ig = wb.create_sheet(title=ignored_sheet_name)
    _write_header(ws_ig, IGNORED_HEADERS, _IGNORED_HEADER_FILL)
    for r in ignored:
        ws_ig.append([
            r.pam_satir, r.username, r.ip_address, r.hostname, r.os, r.safe_name,
            r.domain, r.ignored_reason,
        ])
    ws_ig.freeze_panes = "A2"
    ws_ig.auto_filter.ref = f"A1:{get_column_letter(len(IGNORED_HEADERS))}{ws_ig.max_row}"
    _autosize(ws_ig)

    wb.save(output_file)
    log.info(
        "working_output yazıldı: %s ('%s'=%d satır, '%s'=%d satır).",
        output_file, sheet_name, len(results), ignored_sheet_name, len(ignored),
    )
