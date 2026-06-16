# -*- coding: utf-8 -*-
"""
Excel okuma katmanı.

Sütunlar POZİSYONA göre değil, BAŞLIK ADINA göre okunur. Böylece excel'e
fazladan kolon eklense (PamEnvanter'da olduğu gibi) ya da sıra değişse bile
kod çalışmaya devam eder.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from common.logging_setup import get_logger
from correlation.models import OsRecord, PamRow

log = get_logger("excel_reader")


def _clean(value) -> Optional[str]:
    """Hücre değerini normalize eder: None -> None, sayılar -> str, trim."""
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _pick_sheet(wb: openpyxl.Workbook, sheet_name: Optional[str], file: Path):
    """İstenen sheet'i, yoksa ilk sheet'i döndürür."""
    if sheet_name and sheet_name in wb.sheetnames:
        log.debug("[%s] '%s' sayfası kullanılıyor.", file.name, sheet_name)
        return wb[sheet_name]
    ws = wb.worksheets[0]
    log.debug("[%s] ilk sayfa kullanılıyor: '%s'.", file.name, ws.title)
    return ws


def _build_header_map(ws) -> Dict[str, int]:
    """
    İlk satırı header kabul eder; {kucuk_harf_baslik: sutun_index} döndürür.
    Boşluklar kırpılır, karşılaştırma büyük/küçük harf duyarsızdır.
    """
    header_map: Dict[str, int] = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, cell in enumerate(first_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key and key not in header_map:
            header_map[key] = idx
    return header_map


def _require_columns(header_map: Dict[str, int], required: List[str], file: Path) -> None:
    """Zorunlu kolonların varlığını doğrular; eksikse anlaşılır hata fırlatır."""
    missing = [c for c in required if c.lower() not in header_map]
    if missing:
        raise ValueError(
            f"'{file.name}' dosyasında zorunlu kolon(lar) bulunamadı: {missing}. "
            f"Mevcut kolonlar: {list(header_map.keys())}"
        )


def read_os_envanter(file: Path, sheet_name: Optional[str] = None) -> List[OsRecord]:
    """
    OsEnvanter.xlsx -> OsRecord listesi.

    Beklenen kolonlar: Hostname, IP Address, OS, Domain
    """
    log.info("OsEnvanter okunuyor: %s", file)
    if not file.exists():
        raise FileNotFoundError(f"OsEnvanter dosyası bulunamadı: {file}")

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb, sheet_name, file)
        header = _build_header_map(ws)
        _require_columns(header, ["hostname", "ip address", "os", "domain"], file)

        c_host = header["hostname"]
        c_ip = header["ip address"]
        c_os = header["os"]
        c_domain = header["domain"]

        records: List[OsRecord] = []
        for excel_row, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            hostname = _clean(_get(row, c_host))
            ip = _clean(_get(row, c_ip))
            os_val = _clean(_get(row, c_os))
            domain = _clean(_get(row, c_domain))

            # Tamamen boş satırları atla.
            if not any([hostname, ip, os_val, domain]):
                continue

            records.append(
                OsRecord(
                    hostname=hostname,
                    ip_address=ip,
                    os=os_val,
                    domain=domain,
                    source_row=excel_row,
                )
            )
        log.info("OsEnvanter: %d anlamlı satır okundu.", len(records))
        return records
    finally:
        wb.close()


def read_pam_envanter(file: Path, sheet_name: Optional[str] = None) -> List[PamRow]:
    """
    PamEnvanter.xlsx -> PamRow listesi.

    Part 1 için gerekli kolonlar: userName, safeName, remoteMachines
    (Dosyada fazladan kolonlar olabilir; yok sayılır.)
    """
    log.info("PamEnvanter okunuyor: %s", file)
    if not file.exists():
        raise FileNotFoundError(f"PamEnvanter dosyası bulunamadı: {file}")

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb, sheet_name, file)
        header = _build_header_map(ws)
        _require_columns(header, ["username", "safename", "remotemachines"], file)

        c_user = header["username"]
        c_safe = header["safename"]
        c_remote = header["remotemachines"]
        # 'type' kolonu opsiyonel (yoksa hepsi AD/servis sayılır).
        c_type = header.get("type")

        rows: List[PamRow] = []
        pam_satir = 0  # Sadece veri satırlarını sayar (header hariç).
        for excel_row, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            username = _clean(_get(row, c_user))
            safe_name = _clean(_get(row, c_safe))
            remote = _clean(_get(row, c_remote))
            acc_type = _clean(_get(row, c_type)) if c_type is not None else None

            # Tamamen boş satırları atla.
            if not any([username, safe_name, remote]):
                continue

            pam_satir += 1
            rows.append(
                PamRow(
                    pam_satir=pam_satir,
                    username=username,
                    safe_name=safe_name,
                    remote_machines=remote,
                    source_row=excel_row,
                    account_type=acc_type,
                )
            )
        log.info("PamEnvanter: %d anlamlı satır okundu.", len(rows))
        return rows
    finally:
        wb.close()


def _get(row: tuple, index: int):
    """Satır tuple'ından güvenli erişim (index taşarsa None)."""
    return row[index] if index < len(row) else None
