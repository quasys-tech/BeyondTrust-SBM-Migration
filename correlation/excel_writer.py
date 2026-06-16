# -*- coding: utf-8 -*-
"""
Working.xlsx üreten yazma katmanı.

İki sayfa oluşturur:
  * 'Working'      -> korele başarılı satırlar
  * 'Ignored Rows' -> işlenemeyen satırlar + sebep (INFO)

Başlıklar kalın, filtre açık, satır donduruldu ve kolon genişlikleri ayarlı.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from common.logging_setup import get_logger
from correlation.models import IgnoredRow, WorkingRow

log = get_logger("excel_writer")

# Başlık satırı görünümü
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_IGNORED_HEADER_FILL = PatternFill("solid", fgColor="C0504D")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

WORKING_HEADERS = [
    "PamEnvanterSatır",
    "username",
    "ip address",
    "hostname",
    "OS",
    "safe name",
    "domain",
    "type",
]

IGNORED_HEADERS = [
    "PamEnvanterSatır",
    "Username",
    "Safe Name",
    "Ip Address",
    "Hostname",
    "Domain",
    "OS",
    "INFO",
]


def _style_header(ws: Worksheet, headers: List[str], fill: PatternFill) -> None:
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = _HEADER_ALIGN


def _autosize(ws: Worksheet, headers: List[str], rows: List[list]) -> None:
    """Kolon genişliklerini içeriğe göre (makul sınırlar içinde) ayarlar."""
    for col_idx, title in enumerate(headers, start=1):
        max_len = len(str(title))
        for row in rows:
            if col_idx - 1 < len(row):
                value = row[col_idx - 1]
                if value is not None:
                    max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 10), 60
        )


def _write_sheet(
    ws: Worksheet, headers: List[str], data_rows: List[list], header_fill: PatternFill
) -> None:
    _style_header(ws, headers, header_fill)
    for row in data_rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    if data_rows or headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(data_rows) + 1}"
    _autosize(ws, headers, data_rows)


def write_working_file(
    output_file: Path,
    working_rows: List[WorkingRow],
    ignored_rows: List[IgnoredRow],
    working_sheet_name: str = "Working",
    ignored_sheet_name: str = "Ignored Rows",
) -> None:
    """Working.xlsx dosyasını yazar (varsa üzerine yazar)."""
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # 1) Working sayfası (varsayılan sayfayı yeniden adlandır)
    ws_working = wb.active
    ws_working.title = working_sheet_name
    _write_sheet(
        ws_working,
        WORKING_HEADERS,
        [r.as_excel_row() for r in working_rows],
        _HEADER_FILL,
    )

    # 2) Ignored Rows sayfası
    ws_ignored = wb.create_sheet(title=ignored_sheet_name)
    _write_sheet(
        ws_ignored,
        IGNORED_HEADERS,
        [r.as_excel_row() for r in ignored_rows],
        _IGNORED_HEADER_FILL,
    )

    wb.save(output_file)
    log.info(
        "Working dosyasi yazildi: %s ('%s'=%d satir, '%s'=%d satir).",
        output_file,
        working_sheet_name,
        len(working_rows),
        ignored_sheet_name,
        len(ignored_rows),
    )
