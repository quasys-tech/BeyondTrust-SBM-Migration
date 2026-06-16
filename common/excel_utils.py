# -*- coding: utf-8 -*-
"""Ortak Excel yardımcıları (paylaşılan okuma fonksiyonları)."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from common.logging_setup import get_logger

log = get_logger("excel_utils")


def read_sheet_as_dicts(
    file: Path, sheet_name: Optional[str] = None
) -> List[Dict[str, object]]:
    """
    Bir excel sayfasını {başlık: değer} sözlükleri listesine çevirir.

    İlk satır başlık kabul edilir. Tamamen boş satırlar atlanır.
    """
    file = Path(file)
    if not file.exists():
        raise FileNotFoundError(f"Excel bulunamadı: {file}")

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(headers)]

        records: List[Dict[str, object]] = []
        for row in rows_iter:
            if row is None or not any(v is not None and str(v).strip() for v in row):
                continue
            record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            records.append(record)
        log.info("Excel okundu: %s ('%s') -> %d satır.", file.name, ws.title, len(records))
        return records
    finally:
        wb.close()
